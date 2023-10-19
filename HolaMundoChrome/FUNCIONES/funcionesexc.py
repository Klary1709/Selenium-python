import openpyxl

class FuncionesExc(): 
    def __init__(self, driver):
        #incializa el driver
        self.driver = driver
        
    def getRowCount(file,path,sheetName): 
        """ Esta función cuenta el numero de filas que tenga una hoja de excel 

        Args:
            file (_type_): _description_
            path (str): ruta donde se encuentra la hoja de excel
            sheetName (srt): Nombre de la hoja de excel

        Returns:
            num : un numero maximo de filas 
        """
        worbook=openpyxl.load_workbook(path)
        sheet=worbook[sheetName]
        return (sheet.max_row)
    
    def getColumnCount(file,sheetName): 
        """
        Esta función cuenta el numero de columnas ocupadas en un excel2

        Args:
            file (_type_): _description_
            sheetName (str): Nombre de la hoja de excel

        Returns:
            num: maximo de columnas 
        """
        worbook=openpyxl.load_workbook(file)
        sheet=worbook[sheetName]
        return (sheet.max_column)
      
    def readData(file,path,sheetName,rownum,columno): 
        """Una vez que se cargen la filas y las columnas esta funcion lee los datos 

        Args:
            file (_type_): _description_
            path (str): ruta donde se encuentra la hoja de excel
            sheetName (srt): Nombre de la hoja de excel
            rownum (num): numero de filas
            columno (num): numero de columnas

        Returns:
            regresa el valor de los datos obtenidos 
        """
        workBook=openpyxl.load_workbook(path)
        sheet=workBook[sheetName]
        return sheet.cell(row=rownum,column=columno).value

    def writeData(file,path,sheetName,rownum,columno,data): 
        """
        Se guardan los datos en la hoja de excel 

        Args:
            file (_type_): _description_
            path (str): ruta donde se encuentra la hoja de excel
            sheetName (srt): Nombre de la hoja de excel
            rownum (num): numero de fila 
            columno (num): numero de columna
            data (_type_): una matrix con los datos que se guardaran
        """
        workbook=openpyxl.load_workbook(path)
        sheet=workbook[sheetName]
        sheet.cell(row=rownum,column=columno).value=data
        workbook.save(path)
            
        
        

    